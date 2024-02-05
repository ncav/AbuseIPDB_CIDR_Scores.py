import requests
import ipaddress
import openpyxl
import traceback

def get_abuse_score(ip_address, api_key):
    url = f"https://api.abuseipdb.com/api/v2/check"
    headers = {
        'Accept': 'application/json',
        'Key': api_key
    }
    params = {
        'ipAddress': ip_address,
        'maxAgeInDays': '90'  # Adjust this value as needed
    }
    
    response = requests.get(url, headers=headers, params=params)
    if response.status_code == 200:
        data = response.json()
        score = data.get('data', {}).get('abuseConfidenceScore', 0)
        return score
    else:
        print(f"Error with IP {ip_address}: Received response code {response.status_code}")
        return None

def check_network_range(cidr, api_key, file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'IP Address'
    sheet['B1'] = 'Abuse Score'
    
    network = ipaddress.ip_network(cidr)
    row = 2
    for ip in network.hosts():
        score = get_abuse_score(str(ip), api_key)
        if score is not None:
            print(f"The abuse score of IP address {ip} is: {score}")
            sheet.cell(row=row, column=1, value=str(ip))
            sheet.cell(row=row, column=2, value=score)
            row += 1
        else:
            print(f"Failed to retrieve abuse score for {ip}.")
        
        # Periodically save the workbook after every 10 IPs as an example
        if row % 10 == 0:
            try:
                workbook.save(file_path)
                print(f"Progress saved to {file_path} at row {row}")
            except Exception as e:
                print(f"Failed to save workbook periodically: {e}")
                traceback.print_exc()
    
    # Final save to ensure any remaining data is saved
    try:
        workbook.save(file_path)
        print(f"Final workbook saved successfully to {file_path}")
    except Exception as e:
        print(f"Failed to save final workbook: {e}")
        traceback.print_exc()
# Example usage
api_key = ''  # Replace with your AbuseIPDB API Key
cidr = '' #add in CIDR Range
file_path = r"" #add your windows file path to save in excel sheet
check_network_range(cidr, api_key, file_path)
